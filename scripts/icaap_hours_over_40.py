#!/usr/bin/env python3
"""Report iCAAP month-worked hours over a threshold.

Scans an iCAAP Master workbook and lists every employee/month whose logged
hours exceed a threshold (default 40). Useful for spotting months where an
employee worked more than a normal full week's allotment before those hours
are entered into SAP / TR01.

Reads the "Payroll Tracker" sheet, whose month blocks each begin with an
"Hours" column (row 3 header) under a "<Month> <Year>" label (row 2). Only
numeric cells are considered, so blank / "N/A" cells are ignored.

Usage:
    python scripts/icaap_hours_over_40.py path/to/iCAAP_Master_20262027.xlsx
    python scripts/icaap_hours_over_40.py MASTER.xlsx --threshold 40
    python scripts/icaap_hours_over_40.py MASTER.xlsx --csv > over40.csv

Requires: openpyxl  (pip install openpyxl)
"""

import argparse
import csv
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with: pip install openpyxl")


SHEET_NAME = "Payroll Tracker"
NAME_COL = 2   # column B: employee name
PERN_COL = 3   # column C: PERN
LABEL_ROW = 2  # row with "<Month> <Year>" labels
HEADER_ROW = 3  # row with "Hours" / "Status" / ... headers
DATA_START_ROW = 4


def find_hours_columns(ws):
    """Return [(col_index, month_label)] for every month's Hours column."""
    columns = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(HEADER_ROW, col).value
        if header is None or not str(header).startswith("Hours"):
            continue
        label = ws.cell(LABEL_ROW, col).value
        columns.append((col, str(label).strip() if label else f"col {col}"))
    return columns


def find_over_threshold(ws, threshold):
    """Yield (name, pern, month_label, hours) for hours strictly over threshold."""
    hours_columns = find_hours_columns(ws)
    for row in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(row, NAME_COL).value
        pern = ws.cell(row, PERN_COL).value
        if name is None:
            continue
        for col, month_label in hours_columns:
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if value > threshold:
                    yield (str(name).strip(), pern, month_label, value)


def main():
    parser = argparse.ArgumentParser(
        description="List iCAAP month-worked hours over a threshold (default 40)."
    )
    parser.add_argument("workbook", help="Path to the iCAAP Master .xlsx file")
    parser.add_argument(
        "-t", "--threshold", type=float, default=40,
        help="Report hours strictly greater than this value (default: 40)",
    )
    parser.add_argument(
        "--csv", action="store_true",
        help="Emit results as CSV instead of a formatted table",
    )
    args = parser.parse_args()

    try:
        wb = openpyxl.load_workbook(args.workbook, data_only=True)
    except FileNotFoundError:
        sys.exit(f"Workbook not found: {args.workbook}")

    if SHEET_NAME not in wb.sheetnames:
        sys.exit(
            f'Sheet "{SHEET_NAME}" not found. Available sheets: '
            + ", ".join(wb.sheetnames)
        )

    ws = wb[SHEET_NAME]
    results = list(find_over_threshold(ws, args.threshold))
    # Sort by month order (columns are already chronological), then name.
    # find_over_threshold already yields in row/column order; keep it stable.

    if args.csv:
        writer = csv.writer(sys.stdout)
        writer.writerow(["Name", "PERN", "Month", "Hours"])
        for name, pern, month, hours in results:
            writer.writerow([name, pern, month, _fmt(hours)])
        return

    threshold_str = _fmt(args.threshold)
    if not results:
        print(f"No month-worked hours over {threshold_str} found.")
        return

    name_w = max(len("Name"), max(len(r[0]) for r in results))
    month_w = max(len("Month"), max(len(r[2]) for r in results))
    print(f"iCAAP month-worked hours over {threshold_str}:\n")
    print(f"{'Name':<{name_w}}  {'Month':<{month_w}}  Hours")
    print(f"{'-' * name_w}  {'-' * month_w}  -----")
    for name, _pern, month, hours in results:
        print(f"{name:<{name_w}}  {month:<{month_w}}  {_fmt(hours)}")
    print(f"\n{len(results)} entr{'y' if len(results) == 1 else 'ies'} over {threshold_str}.")


def _fmt(value):
    """Render whole numbers without a trailing .0."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


if __name__ == "__main__":
    main()
